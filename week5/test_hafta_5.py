from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from constants import globalConstants as glob
import pytest
from pathlib import Path
from datetime import date
import openpyxl


class Test_Swag_labs:

    # her testten önce çağırılır
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(glob.URL)
        # günün tarihini al bu tarih ile bir klasör var mı kontrol et yoksa oluştur
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)

    # her testten sonra çağırılır
    def teardown_method(self):
        self.driver.quit()

    # sayfanın yüklendiğini kontrol eden fonksiyon
    def waitForUrlToBe(self, url, timeout=5):
        WebDriverWait(self.driver, timeout).until(ec.url_to_be(url))

    # elementin görünür olduğunu kontrol eden fonksiyon
    def waitForElementVisible(self, locator, timeout=5):
        WebDriverWait(self.driver, timeout).until(
            ec.visibility_of_element_located(locator))

    # excel dosyasyından veri al
    def getData(sheetName):
        # veriyi al
        excelFile = openpyxl.load_workbook("data/userLoginInformation.xlsx")
        selectedSheet = excelFile[sheetName]
        totalRows = selectedSheet.max_row
        data = []
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            tupleData = (username, password)
            data.append(tupleData)
        return data

    # boş kullanıcı adı ve şifre gönderilme durumunu test et
    def test_empty_username_pasword(self):
        # login butonuna tıkla
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # hata mesajını bul
        self.waitForElementVisible((By.CLASS_NAME, "error-button"))
        errMessage = self.driver.find_element(
            By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_empty_username_pasword.png")
        # hata mesajını kontrol et
        assert errMessage.text == glob.emptyUsernamePasword

    # boş şifre gönderilme durumunu test et
    def test_empty_pasword(self):
        # kullanıcı adır gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        inputUsername.send_keys("user")
        # login butonuna tıkla
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # hata mesajını bul
        errMessage = self.driver.find_element(
            By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_empty_pasword.png")
        # hata mesajını kontrol et
        assert errMessage.text == glob.emptyPasword

    # teste parametreleri kaynak dosyasından gönder
    @pytest.mark.parametrize("username,password", getData("locked_user"))
    def test_locked_user(self, username, password):
        # kullanıcı adı ve şifre gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        inputPassword = self.driver.find_element(By.ID, "password")
        inputUsername.send_keys(username)
        inputPassword.send_keys(password)
        # giriş butonuna tıkla
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # hata mesajını bul
        errMessage = self.driver.find_element(
            By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_locked_user-{username}-{password}.png")
        # hata mesajını kontrol et
        assert errMessage.text == glob.lockedUser

    # boş girdiler ile giriş yapınca çıkan x ikonlarını test et
    def test_icon_x(self):
        # giriş butonuna tıkla
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # hata ikonlarını bul
        self.waitForElementVisible((By.CLASS_NAME, "error_icon"))
        icons = self.driver.find_elements(By.CLASS_NAME, "error_icon")
        # hata ikonlarının sayısını kontrol et
        numOfIcon = len(icons)
        assert numOfIcon == 2
        # hata ikonlarını kapat
        self.waitForElementVisible((By.CLASS_NAME, "error-button"))
        errMessageClose = self.driver.find_element(
            By.CLASS_NAME, "error-button")
        errMessageClose.click()
        # ikonları tekrar ara
        icons = self.driver.find_elements(By.CLASS_NAME, "error_icon")
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_icon_x.png")
        # hata ikonlarının sayısını tekrar kontrol et
        if not icons:
            assert True
        else:
            assert False

    # teste parametreleri kaynak dosyasından gönder
    @pytest.mark.parametrize("username,password", getData("standard_user"))
    def test_standard_user(self, username, password):
        # kullanıcı adı şifre gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        inputPassword = self.driver.find_element(By.ID, "password")
        inputUsername.send_keys(username)
        inputPassword.send_keys(password)
        # giriş butonuna tıkla
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        self.waitForUrlToBe("https://www.saucedemo.com/inventory.html")
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_standard_user-{username}-{password}.png")
        # sayfa adresini kontrol et
        assert "https://www.saucedemo.com/inventory.html" in self.driver.current_url

    # teste parametreleri kaynak dosyasından gönder
    # burada birden fazla kullanıcı için test yapılıyor -> bakınız: data/userLoginInformation.xlsx
    @pytest.mark.parametrize("username,password", getData("multi_user"))
    def test_number_of_product(self, username, password):
        # kullanıcı adı şifre gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        inputPassword = self.driver.find_element(By.ID, "password")
        inputUsername.send_keys(username)
        inputPassword.send_keys(password)
        # giriş butonuna tıkla
        self.waitForElementVisible((By.ID, "login-button"), 10)
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_number_of_product-{username}-{password}.png")
        # ürün sayısını bul ve kontrol et
        products = self.driver.find_elements(By.CLASS_NAME, "inventory_item")
        assert len(products) == 6

    # teste parametreleri kaynak dosyasından gönder
    @pytest.mark.parametrize("username,password", getData("standard_user"))
    def test_shopping_add_badge(self, username, password):
        # kullanıcı adı şifre gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        inputPassword = self.driver.find_element(By.ID, "password")
        inputUsername.send_keys("standard_user")
        inputPassword.send_keys("secret_sauce")
        # giriş butonuna tıkla
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # ürünleri bul ve listeye al
        self.waitForElementVisible((By.CLASS_NAME, "btn_inventory"))
        products = self.driver.find_elements(By.CLASS_NAME, "btn_inventory")
        # ürünleri ekle
        products[0].click()
        products[1].click()
        self.waitForElementVisible((By.CLASS_NAME, "shopping_cart_badge"))
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_shopping_add_badge-{username}-{password}.png")
        # sepete ürün eklenme rozetini kontrol et
        badge = self.driver.find_element(By.CLASS_NAME, "shopping_cart_badge")
        assert badge.is_displayed()

    # teste parametreleri kaynak dosyasından gönder
    @pytest.mark.parametrize("username,password", getData("problem_user"))
    def test_problem_user_checkout(self, username, password):
        # kullanıcı adı şifre gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        inputPassword = self.driver.find_element(By.ID, "password")
        inputUsername.send_keys(username)
        inputPassword.send_keys(password)
        # giriş yap
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # sepete git
        self.waitForElementVisible((By.CLASS_NAME, "shopping_cart_link"))
        shoppingCart = self.driver.find_element(
            By.CLASS_NAME, "shopping_cart_link")
        shoppingCart.click()
        # kontrole git
        self.waitForElementVisible((By.ID, "checkout"))
        btnCheckout = self.driver.find_element(By.ID, "checkout")
        btnCheckout.click()
        # isim soyisim gir
        first_name = self.driver.find_element(By.ID, "first-name")
        last_name = self.driver.find_element(By.ID, "last-name")
        first_name.send_keys(username)
        last_name.send_keys(password)
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_problem_user_checkout-{username}-{password}.png")
        # soyisim girildiğinde isim kısmında görünüyor mu kontrol et
        assert first_name.get_attribute("value") == "e"
        assert last_name.get_attribute("value") == ""

    # teste parametreleri kaynak dosyasından gönder
    @pytest.mark.parametrize("username,password", getData("standard_user"))
    def test_full_shopping(self, username, password):
        # kullanıcı adı şifre gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        inputPassword = self.driver.find_element(By.ID, "password")
        inputUsername.send_keys(username)
        inputPassword.send_keys(password)
        # giriş yap
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # ürün sayfasının yüklendiğini kontrol et
        self.waitForUrlToBe("https://www.saucedemo.com/inventory.html")
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_full_shopping_product_page-{username}-{password}.png")
        assert "https://www.saucedemo.com/inventory.html" in self.driver.current_url
        # en pahalı 2 ürünü sepete ekle
        # İlk olarak, tüm "add to cart" düğmelerini toplar. Daha sonra, sorted() fonksiyonu kullanılarak, fiyatları yüksekten düşüğe doğru sıralanır. Bu sıralama işlemi, key parametresine verilen lambda fonksiyonu ile gerçekleştirilir. Bu lambda fonksiyonu, her düğmenin üst düzey elementindeki fiyat etiketini alır ve ona göre sıralama yapar (reverse ifadesinin varsayılanı false dır diziyi tersine çevirme anlamına gelir) Sıralama tamamlandıktan sonra, ilk iki düğme (sorted_buttons[:2]) sepete eklenir
        self.waitForElementVisible(
            (By.CLASS_NAME, "pricebar"))
        itemPrices = self.driver.find_elements(
            By.CLASS_NAME, "pricebar")
        sortedPrices = sorted(itemPrices, key=lambda item: float(self.driver.find_element(
            By.CLASS_NAME, "inventory_item_price").text.replace("$", "").split('\n')[0]), reverse=True)
        for button in sortedPrices[:2]:
            button.click()
        # sepete git
        self.waitForElementVisible((By.CLASS_NAME, "shopping_cart_link"))
        shoppingCart = self.driver.find_element(
            By.CLASS_NAME, "shopping_cart_link")
        shoppingCart.click()
        # sepet sayfasının yüklenmesini kontrol et
        self.waitForUrlToBe("https://www.saucedemo.com/cart.html")
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_full_shopping_cart_page-{username}-{password}.png")
        assert "https://www.saucedemo.com/cart.html" in self.driver.current_url
        # checkout sayfasına git
        self.waitForElementVisible((By.ID, "checkout"))
        btnCheckout = self.driver.find_element(By.ID, "checkout")
        btnCheckout.click()
        # isim soyisim posta kodu gir
        self.waitForElementVisible((By.ID, "first-name"))
        firstName = self.driver.find_element(By.ID, "first-name")
        self.waitForElementVisible((By.ID, "last-name"))
        lastName = self.driver.find_element(By.ID, "last-name")
        self.waitForElementVisible((By.ID, "postal-code"))
        postalCode = self.driver.find_element(By.ID, "postal-code")
        firstName.send_keys("firstname")
        lastName.send_keys("lastname")
        postalCode.send_keys("1")
        # continue düğmesine tıkla
        self.waitForElementVisible((By.ID, "continue"))
        continue_button = self.driver.find_element(By.ID, "continue")
        continue_button.click()
        # ödeme sayfasının yüklenmesini kontrol et
        self.waitForUrlToBe(
            "https://www.saucedemo.com/checkout-step-two.html")
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_full_shopping_pay_page-{username}-{password}.png")
        assert "https://www.saucedemo.com/checkout-step-two.html" in self.driver.current_url
        # ürünlerin  fiyat toplamı total fiyata eşit mi
        # ürün fiyatlarını topla
        self.waitForElementVisible((By.CLASS_NAME, "inventory_item_price"))
        prices = self.driver.find_elements(
            By.CLASS_NAME, "inventory_item_price")
        sumOfPrices = 0.0
        for price in prices:
            sumOfPrices += float(price.text.split('$')[-1])
        # vergiyi fiyata ekle
        self.waitForElementVisible((By.CLASS_NAME, "summary_tax_label"))
        taxPrice = self.driver.find_element(
            By.CLASS_NAME, "summary_tax_label").text.split("$")[1]
        sumOfPrices += float(taxPrice)
        # total fiyatı bul
        self.waitForElementVisible((By.CLASS_NAME, "summary_total_label"))
        total = self.driver.find_element(
            By.CLASS_NAME, "summary_total_label").text.split('$')[-1]
        assert sumOfPrices == float(total)
        # finish butonuna tıkla
        self.waitForElementVisible((By.ID, "finish"))
        btnLogin = self.driver.find_element(By.ID, "finish")
        btnLogin.click()
        # ödeme tamamlandı sayfasının mesajını kontrol et(Thank you for your order!)
        self.waitForElementVisible((By.CLASS_NAME, "complete-header"))
        completeHeader = self.driver.find_element(
            By.CLASS_NAME, "complete-header")
        # ekran görüntüsü al
        self.driver.save_screenshot(
            f"{self.folderPath}/test_full_shopping_complete_page-{username}-{password}.png")
        assert completeHeader.text == glob.completeHeader
        # back home düğmesine tıkla
        self.waitForElementVisible((By.ID, "back-to-products"))
        btnLogin = self.driver.find_element(By.ID, "back-to-products")
        btnLogin.click()
        # ürün sayfasının yüklendiğini kontrol et
        self.waitForUrlToBe("https://www.saucedemo.com/inventory.html")
        assert "https://www.saucedemo.com/inventory.html" in self.driver.current_url
