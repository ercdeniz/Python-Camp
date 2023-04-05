from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from constants import globalConstants as glob
import pytest
from pathlib import Path
from datetime import date
import openpyxl
import requests


class Test_Swag_labs_2:
    # her testten önce çağırılır
    def setup_method(self):
        self.driver = webdriver.Chrome(
            ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(glob.URL)
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)

    # her testten sonra çağırılır
    def teardown_method(self):
        self.driver.quit()

    # sayfanın yüklendiğini kontrol eden fonksiyon
    def waitForUrlToBe(self, url, timeout=5):
        WebDriverWait(self.driver, timeout).until(ec.url_to_be(url))

    # elementin görünür olduğunu kontrol eden fonksiyon
    def waitForElementVisible(self, locator, timeout=5):
        WebDriverWait(self.driver, timeout).until(
            ec.visibility_of_element_located(locator))

     # excel dosyasyından veri al
    def getData(sheetName):
        # veriyi al
        excelFile = openpyxl.load_workbook("data/userLoginInformation.xlsx")
        selectedSheet = excelFile[sheetName]
        totalRows = selectedSheet.max_row
        data = []
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            tupleData = (username, password)
            data.append(tupleData)
        return data

    # ana sayfanın yüklenmesini kontrol et
    def test_home_page_load(self):
        self.waitForUrlToBe(glob.URL)
        assert "Swag Labs" in self.driver.title

    # teste parametreleri kaynak dosyasından gönder
    @pytest.mark.parametrize("username,password", getData("multi_user"))
    # doğru giriş yapıldığında doğru sayfa yükleniyor mu
    def test_login_successful(self, username, password):
        # kullanıcı adı şifre gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        inputPassword = self.driver.find_element(By.ID, "password")
        inputUsername.send_keys(username)
        inputPassword.send_keys(password)
        # giriş yap
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # giriş sayfasının yüklendiğini kontrol et
        self.waitForUrlToBe(glob.productPageURL)
        assert glob.productPageURL in self.driver.current_url

    # teste parametreleri kaynak dosyasından gönder
    # burada birden fazla kullanıcı için test yapılıyor -> bakınız: data/userLoginInformation.xlsx
    @pytest.mark.parametrize("username,password", getData("multi_user"))
    # sayfanın yüklenme hızını kontrol et
    def test_page_load_time(self, username, password):
        # kullanıcı adı şifre gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        inputPassword = self.driver.find_element(By.ID, "password")
        inputUsername.send_keys(username)
        inputPassword.send_keys(password)
        # giriş butonuna tıkla
        self.waitForElementVisible((By.ID, "login-button"))
        # giriş butonuna tıklamadan önce sayacı başlatıyoruz
        start_time = time.time()
        login_button = self.driver.find_element(By.ID, "login-button")
        login_button.click()
        # tıklama işlemi yapıldıktan sonra sayacı durduruyoruz
        end_time = time.time()
        # farkı buluyoruz
        total_time = end_time - start_time
        # performans_glitch_user giriş sırasında çok geç bağlandığı için onu farklı test ettim
        if username != "performance_glitch_user":
            assert total_time < 3.0
        else:
            assert total_time < 10.0

    # teste parametreleri kaynak dosyasından gönder
    @pytest.mark.parametrize("username,password", getData("standard_user"))
    def test_payment_failure(self, username, password):
        # kullanıcı adı şifre gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        inputPassword = self.driver.find_element(By.ID, "password")
        inputUsername.send_keys(username)
        inputPassword.send_keys(password)
        # giriş yap
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # sepete git
        self.waitForElementVisible((By.CLASS_NAME, "shopping_cart_link"))
        shoppingCart = self.driver.find_element(
            By.CLASS_NAME, "shopping_cart_link")
        shoppingCart.click()
        self.waitForElementVisible((By.ID, "checkout"))
        btnCheckout = self.driver.find_element(By.ID, "checkout")
        btnCheckout.click()
        # isim soyisim posta kodu gir
        self.waitForElementVisible((By.ID, "first-name"))
        firstName = self.driver.find_element(By.ID, "first-name")
        self.waitForElementVisible((By.ID, "last-name"))
        lastName = self.driver.find_element(By.ID, "last-name")
        self.waitForElementVisible((By.ID, "postal-code"))
        postalCode = self.driver.find_element(By.ID, "postal-code")
        firstName.send_keys("")
        lastName.send_keys("")
        postalCode.send_keys("")
        # continue düğmesine tıkla
        self.waitForElementVisible((By.ID, "continue"))
        continue_button = self.driver.find_element(By.ID, "continue")
        continue_button.click()
        # error mesajını ara
        self.waitForElementVisible(
            (By.XPATH, "//*[@id = 'checkout_info_container']/div/form/div[1]/div[4]/h3"))
        errorMessage = self.driver.find_element(
            By.XPATH, "//*[@id = 'checkout_info_container']/div/form/div[1]/div[4]/h3")
        # error mesajını kontrol et
        assert errorMessage.text == glob.checkoutErrorMessage

    @pytest.mark.parametrize("username,password", getData("standard_user"))
    def test_add_and_remove_items_to_cart(self, username, password):
        # kullanıcı adı şifre gir
        self.waitForElementVisible((By.ID, "user-name"))
        inputUsername = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        inputPassword = self.driver.find_element(By.ID, "password")
        inputUsername.send_keys(username)
        inputPassword.send_keys(password)
        # giriş yap
        self.waitForElementVisible((By.ID, "login-button"))
        btnLogin = self.driver.find_element(By.ID, "login-button")
        btnLogin.click()
        # ürün sayfasının yüklendiğini kontrol et
        self.waitForUrlToBe(glob.productPageURL)
        assert glob.productPageURL in self.driver.current_url
        # en ucuz 2 ürünü sepete ekle
        # İlk olarak, tüm "add to cart" düğmelerini toplar. Daha sonra, sorted() fonksiyonu kullanılarak, fiyatları düşükten yükseğe doğru sıralanır. Bu sıralama işlemi, key parametresine verilen lambda fonksiyonu ile gerçekleştirilir. Bu lambda fonksiyonu, her düğmenin üst düzey elementindeki fiyat etiketini alır ve ona göre sıralama yapar. Sıralama tamamlandıktan sonra, ilk iki düğme (sorted_buttons[:2]) sepete eklenir
        self.waitForElementVisible(
            (By.CLASS_NAME, "pricebar"))
        itemPrices = self.driver.find_elements(
            By.CLASS_NAME, "pricebar")
        sortedPrices = sorted(itemPrices, key=lambda item: float(self.driver.find_element(
            By.CLASS_NAME, "inventory_item_price").text.replace("$", "").split('\n')[0]))
        for button in sortedPrices[:2]:
            button.click()
        # sepete git
        self.waitForElementVisible((By.CLASS_NAME, "shopping_cart_link"))
        shoppingCart = self.driver.find_element(
            By.CLASS_NAME, "shopping_cart_link")
        shoppingCart.click()
        # sepet sayfasının yüklenmesini kontrol et
        self.waitForUrlToBe(glob.cartPageURL)
        # ürünleri kaldır
        removeButtons = self.driver.find_elements(
            By.CLASS_NAME, "btn btn_secondary")
        for button in removeButtons:
            button.click()
        # ürünlerin kaldırılmasını kontrol et
        itemBar = self.driver.find_element(By.CLASS_NAME, "cart_item")
        assert itemBar.is_displayed()
